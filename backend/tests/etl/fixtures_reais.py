"""Construtores de planilhas sintéticas com as ESTRUTURAS REAIS (Sprint 3).

Reproduzem fielmente os formatos descobertos na engenharia reversa
(docs/DATA_PROFILING.md): matriz de faturamento com 2 linhas de cabeçalho
e rodapé, relatório Supervisor com blocos duplicados, export de checklists
com 8 abas/42 colunas, formulários WeCheck com schema drift e Painel Avert
— com dados 100% fictícios (o repositório é público).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from etl.arquivos import FluxoArquivos

_TIMESTAMP_DETERMINISTICO = datetime(2026, 1, 1, 0, 0, 0)


def _salvar(workbook: Workbook, destino: Path) -> Path:
    workbook.properties.created = _TIMESTAMP_DETERMINISTICO
    workbook.save(destino)
    return destino


def xlsx_base_clientes(
    fluxo: FluxoArquivos,
    nome: str = "base_clientes.xlsx",
    linhas: list[list[Any]] | None = None,
) -> Path:
    """Base de Clientes real: 22 colunas, com pares RCA e 'Nome RCA' repetido."""

    colunas = [
        "Código",
        "CNPJ/CPF",
        "Insc. Est. / Produtor",
        "Cliente",
        "Fantasia",
        "Tipo de Pessoa",
        "Ramo Atividade",
        "Endereço Comercial",
        "Número",
        "Bairro",
        "Nome da Cidade",
        "Estado",
        "CEP",
        "Telefone",
        "RCA 1",
        "Nome RCA",
        "RCA 2",
        "Nome RCA",
        "RCA 3",
        "Nome RCA",
        "RCA4",
        "Data da Última Compra",
    ]
    if linhas is None:
        linhas = [
            [
                "10001",
                "11222333000144",
                "ISENTO",
                "PET SHOP ALFA LTDA",
                "PET ALFA",
                "Jurídica(J)",
                "PET SHOP",
                "RUA A",
                "100",
                "CENTRO",
                "CAMPINAS",
                "SP",
                "13000-000",
                "(19) 3333-4444",
                "77",
                "VENDEDOR SETENTA E SETE",
                None,
                None,
                None,
                None,
                None,
                datetime(2026, 5, 10, 14, 30),
            ],
            [
                "10002",
                "22333444000155",
                None,
                "CLINICA VET BETA",
                None,
                "Jurídica(J)",
                "CLÍNICA VETERINÁRIA",
                "AV B",
                "200",
                "JARDIM",
                "SÃO PAULO",
                "SP",
                "01000-000",
                None,
                "77",
                "VENDEDOR SETENTA E SETE",
                "88",
                "VENDEDOR OITENTA E OITO",
                None,
                None,
                None,
                None,
            ],
            [
                "10003",
                None,
                None,
                "AGRO GAMA PRODUTOR",
                "GAMA",
                "Física(F)",
                "PRODUTOR RURAL",
                None,
                None,
                None,
                "BELO HORIZONTE",
                "MG",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = "BASE DE CLIENTES -  BRASIL"
    aba.append(colunas)
    for linha in linhas:
        aba.append(linha)
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_faturamento_matriz(
    fluxo: FluxoArquivos,
    nome: str = "faturamento_janeiro.xlsx",
    marcas: list[str] | None = None,
    linhas_clientes: list[tuple[str, list[Any]]] | None = None,
    mes_rodape: str = "Janeiro",
    ano_rodape: int = 2026,
) -> Path:
    """Matriz real Cliente×Marca: 2 linhas de cabeçalho, Total e rodapé."""

    marcas = marcas or ["AVERT", "BBPET", "BRINDE"]
    if linhas_clientes is None:
        linhas_clientes = [
            ("10001 - PET SHOP ALFA LTDA", ["396.07", None, "10.00"]),
            ("10002 - CLINICA VET BETA", [None, "1250.50", None]),
            ("10003 - AGRO GAMA PRODUTOR", ["-120.00", "80.00", None]),
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Export"
    aba.append(["Departamento", *marcas, "Total"])
    aba.append(["Cliente", *["Medidas faturamento"] * (len(marcas) + 1)])
    for rotulo, valores in linhas_clientes:
        total = sum(float(v) for v in valores if v is not None)
        aba.append([rotulo, *valores, f"{total:.2f}"])
    aba.append(["Total", *[None] * len(marcas), None])
    aba.append(
        [
            "Filtros aplicados: Nome é Faturamento; Operação é Venda ou Devolução; "
            f"Incluídos (1) {ano_rodape} (Ano) + {mes_rodape} (Mês); NUMNOTA não é 94083"
        ]
    )
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_sb_supervisor(
    fluxo: FluxoArquivos,
    nome: str = "sb_supervisor.xlsx",
    linhas: list[list[Any]] | None = None,
) -> Path:
    """Relatório Supervisor real: dois blocos lado a lado, cabeçalhos repetidos."""

    colunas = [
        "Código",
        "Nome",
        "Área",
        "Visitas Previstas",
        "Previstas Realizadas",
        "Previstas Não Realizadas",
        "Não Prevista Realizadas",
        "Código",
        "Nome Fantasia",
        "Razão Social",
        "Visitas Previstas",
        "Visitas Realizadas",
        "Não Visitas",
        "Porcentagem de Visitas á Realizar",
        "Porcentagem de Visitas Realizadas",
        "Porcentagem de  Não Visitas",
    ]
    if linhas is None:
        linhas = [
            [
                "0343",
                "PROMOTORA UM",
                "RS",
                "19",
                "10",
                "9",
                "2",
                "10001",
                "PET ALFA",
                "PET SHOP ALFA LTDA",
                "2",
                "2",
                "0",
                "0%",
                "100%",
                "0%",
            ],
            [
                "0343",
                "PROMOTORA UM",
                "RS",
                "19",
                "10",
                "9",
                "2",
                "10002",
                None,
                "CLINICA VET BETA",
                "1",
                "0",
                "1",
                "100%",
                "0%",
                "100%",
            ],
            [
                "0777",
                "PROMOTOR DOIS",
                "SP",
                "5",
                "5",
                "0",
                "1",
                "10003",
                "GAMA",
                "AGRO GAMA PRODUTOR",
                "1",
                "1",
                "0",
                "0%",
                "100%",
                "0%",
            ],
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Supervisor"
    aba.append(colunas)
    for linha in linhas:
        aba.append(linha)
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_sb_produtos(
    fluxo: FluxoArquivos,
    nome: str = "sb_produtos.xlsx",
    linhas: list[list[Any]] | None = None,
) -> Path:
    """Detalhe real do SB: 4 abas, dados apenas em `Produtos`."""

    colunas = [
        "VISITA",
        "CÓDIGO",
        "FUNCIONARIO",
        "REGIÃO",
        "COD. CLIENTE",
        "RAZAO SOCIAL",
        "NOME FANTASIA",
        "DATA INICIAL",
        "DATA FINAL",
        "OPERAÇÃO",
        "GRUPO",
        "MARCA",
        "COD. PRODUTO",
        "PRODUTO",
        "VALIDADE",
        "LOTE",
        "ESTOQUE",
        "PREÇO",
        "OBSERVAÇÃO",
    ]
    if linhas is None:
        linhas = [
            [
                "29883",
                "0343",
                "PROMOTORA UM",
                "RS",
                "10001",
                "PET SHOP ALFA LTDA",
                "PET ALFA",
                "03/06/2026",
                "08/06/2026",
                "Operação produtos",
                "BBPET",
                "Sem Marca",
                "3072",
                "PRODUTO TESTE 300MG",
                None,
                None,
                "1",
                "0.00",
                None,
            ],
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Produtos"
    aba.append(colunas)
    for linha in linhas:
        aba.append(linha)
    for vazia in ("Gondola", "ProdutoSimilar", "Tarefas"):
        workbook.create_sheet(vazia)
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_checklist_sb(
    fluxo: FluxoArquivos,
    nome: str = "checklist_junho.xlsx",
    abas: dict[str, dict[str, Any]] | None = None,
) -> Path:
    """Export real de checklists: contexto (com CÓDIGO duplicado) + perguntas wide.

    `abas` mapeia título → {"ck_id": int, "checklist": nome, "linhas": [...]}
    onde cada linha é [cod_promotor, funcionario, uf, cod_cliente, razao,
    fantasia, visita, aplicacao, respostas...].
    """

    perguntas = ["Descrever a visita", "FOTO", "Quem é o RCA responsável pela loja"]
    colunas = [
        " CÓDIGO ",
        " FUNCIONÁRIO ",
        " UF ",
        " CÓDIGO ",
        " RAZÃO SOCIAL ",
        " FANTASIA ",
        " VISITA ",
        " CK_ID ",
        " CHECKLIST ",
        " APLICAÇÃO ",
        *perguntas,
    ]
    if abas is None:
        abas = {
            "Check-list Visita Técnica": {
                "ck_id": 6,
                "checklist": "Check-list Visita Técnica",
                "linhas": [
                    [
                        "0343",
                        "PROMOTORA UM",
                        "RS",
                        "10001",
                        "PET SHOP ALFA LTDA",
                        "PET ALFA",
                        "29883",
                        "05/06/2026 10:30",
                        "Visita produtiva",
                        None,
                        "77",
                    ],
                ],
            },
            "Checklist Trade": {
                "ck_id": 7,
                "checklist": "Checklist Trade",
                "linhas": [
                    [
                        "0777",
                        "PROMOTOR DOIS",
                        "SP",
                        "10002",
                        "CLINICA VET BETA",
                        None,
                        "29884",
                        "06/06/2026 15:00",
                        None,
                        "https://fotos.exemplo/1.jpg",
                        None,
                    ],
                ],
            },
        }
    workbook = Workbook()
    workbook.remove(workbook.active)
    for titulo, config in abas.items():
        aba = workbook.create_sheet(titulo)
        aba.append(colunas)
        for linha in config["linhas"]:
            cod_p, func, uf, cod_c, razao, fant, visita, aplicacao, *respostas = linha
            aba.append(
                [
                    cod_p,
                    func,
                    uf,
                    cod_c,
                    razao,
                    fant,
                    visita,
                    config["ck_id"],
                    config["checklist"],
                    aplicacao,
                    *respostas,
                ]
            )
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_wecheck(
    fluxo: FluxoArquivos,
    nome: str = "wecheck_jan.xlsx",
    perguntas: list[str] | None = None,
    linhas: list[list[Any]] | None = None,
    formulario: str = "Visita de Trade",
) -> Path:
    """Export real WeCheck: contexto + perguntas do formulário (wide)."""

    perguntas = perguntas or [
        "Tire fotos da fachada:",
        "Há produtos em ruptura?",
        "Visão crítica da visita:",
    ]
    colunas = [
        "Formulário",
        "Data / Hora do Item",
        "Local",
        "Endereço",
        "Cidade",
        "Estado",
        "Autor",
        "Tarefa",
        "Descrição",
        *perguntas,
    ]
    if linhas is None:
        linhas = [
            [
                formulario,
                "15/01/2026 09:12",
                "AGROPET EXEMPLO",
                "RUA X, 1",
                "PORTO ALEGRE",
                "RS",
                "PROMOTORA WECHECK UM",
                "Visita mensal",
                None,
                "https://fotos.exemplo/f.jpg",
                "Não",
                "Loja organizada",
            ],
            [
                formulario,
                "16/01/2026 14:40",
                "PETSHOP MODELO",
                None,
                "CANOAS",
                "RS",
                "PROMOTORA WECHECK UM",
                None,
                None,
                None,
                "Sim",
                "Faltou produto Y",
            ],
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = f"3 - {formulario.lower().replace(' ', '-')}..."
    aba.append(colunas)
    for linha in linhas:
        aba.append(linha)
    return _salvar(workbook, fluxo.incoming / nome)


def xlsx_painel_avert(
    fluxo: FluxoArquivos,
    nome: str = "painel_avert.xlsx",
    linhas: list[list[Any]] | None = None,
) -> Path:
    """Painel Trade Avert real: carteira por CNPJ com CONSULTOR (promotora)."""

    colunas = [
        "CNPJ",
        "COMPRA 2025",
        "COMPRA 2026",
        "CRESC",
        "UF",
        "ÁREA",
        "REGIONAL",
        "DISTRIBUIDOR",
        "COORDENADOR",
        "CONSULTOR",
        "VENDEDOR",
        "GRUPO ECONÔMICO",
        "NOME FANTASIA",
        "RAZÃO SOCIAL",
        "SEGMENTO",
        "OBS:",
    ]
    if linhas is None:
        linhas = [
            [
                "11.222.333/0001-44",
                None,
                None,
                None,
                "RS",
                "SUL",
                "REGIONAL 1",
                "DISTRIBUIDOR X",
                "COORD A",
                "PROMOTORA WECHECK UM",
                "VENDEDOR Z",
                "GRUPO ALFA",
                "PET ALFA",
                "PET SHOP ALFA LTDA",
                "PET SHOP",
                None,
            ],
            [
                "99.888.777/0001-66",
                None,
                None,
                None,
                "RS",
                "SUL",
                "REGIONAL 1",
                "DISTRIBUIDOR X",
                "COORD A",
                "PROMOTORA WECHECK UM",
                "VENDEDOR Z",
                None,
                "LOJA SEM CADASTRO",
                "LOJA SEM CADASTRO LTDA",
                "AGRO",
                None,
            ],
        ]
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Planilha1"
    aba.append(colunas)
    for linha in linhas:
        aba.append(linha)
    return _salvar(workbook, fluxo.incoming / nome)
