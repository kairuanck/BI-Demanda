"""Hash de conteúdo lógico de planilhas (Sprint 3, docs/DECISIONS.md, 11.2).

Os exports reais chegam em dezenas de cópias fisicamente distintas
(metadados internos do xlsx variam a cada export) porém idênticas em
conteúdo de células — o SHA-256 binário não detecta essas duplicatas.
Este módulo calcula um SHA-256 determinístico sobre os valores das
células, aba a aba, linha a linha, independente de metadados, formatação
e timestamps internos do arquivo.
"""

from __future__ import annotations

import hashlib
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_SEP_CELULA = b"\x1f"  # unit separator
_SEP_LINHA = b"\x1e"  # record separator
_SEP_ABA = b"\x1d"  # group separator


def _celula_canonica(valor: Any) -> bytes:
    """Representação canônica e estável de um valor de célula."""

    if valor is None:
        return b""
    if isinstance(valor, datetime | date | time):
        return valor.isoformat().encode("utf-8")
    if isinstance(valor, float) and valor.is_integer():
        # 123.0 e 123 devem produzir o mesmo hash (o Excel alterna os dois)
        return str(int(valor)).encode("utf-8")
    return str(valor).encode("utf-8")


def calcular_hash_conteudo(caminho: Path) -> str:
    """SHA-256 do conteúdo lógico de todas as abas do arquivo xlsx."""

    sha256 = hashlib.sha256()
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        for aba in workbook.worksheets:
            sha256.update(aba.title.encode("utf-8"))
            sha256.update(_SEP_ABA)
            for linha in aba.iter_rows(values_only=True):
                for valor in linha:
                    sha256.update(_celula_canonica(valor))
                    sha256.update(_SEP_CELULA)
                sha256.update(_SEP_LINHA)
    finally:
        workbook.close()
    return sha256.hexdigest()
