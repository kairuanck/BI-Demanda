"""Leitura bruta multi-aba para os conectores reais (Sprint 3).

Os exports reais violam as premissas do leitor documental (`ler_excel`):
múltiplas abas, cabeçalhos duplicados (dois blocos "Código" lado a lado),
matrizes pivotadas e rodapés. Este módulo entrega as células cruas por
aba, preservando a POSIÇÃO das colunas — cada conector aplica seu próprio
mapeamento adaptativo por nome normalizado (docs/DECISIONS.md, 11.3).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from etl.readers.excel_reader import normalizar_nome_coluna


@dataclass(frozen=True)
class AbaBruta:
    """Uma aba lida integralmente: título e matriz de células (linha 1 = índice 0)."""

    titulo: str
    linhas: list[tuple[Any, ...]]

    @property
    def vazia(self) -> bool:
        return len(self.linhas) <= 1  # só cabeçalho (ou nada)

    def cabecalho(self) -> list[str | None]:
        """Nomes originais da primeira linha (com trim), preservando posição."""

        if not self.linhas:
            return []
        return [str(c).strip() if c is not None else None for c in self.linhas[0]]

    def cabecalho_normalizado(self) -> list[str | None]:
        return [normalizar_nome_coluna(c) if c is not None else None for c in self.cabecalho()]


def ler_abas(caminho: Path) -> list[AbaBruta]:
    """Lê todas as abas do xlsx como valores crus (sem fórmulas/formatação)."""

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        abas: list[AbaBruta] = []
        for aba in workbook.worksheets:
            linhas = [tuple(linha) for linha in aba.iter_rows(values_only=True)]
            # remove linhas totalmente vazias no fim (comuns em exports)
            while linhas and all(v is None for v in linhas[-1]):
                linhas.pop()
            abas.append(AbaBruta(titulo=aba.title, linhas=linhas))
        return abas
    finally:
        workbook.close()


def indices_por_nome(cabecalho_normalizado: list[str | None]) -> dict[str, list[int]]:
    """Mapeia nome normalizado → posições (suporta cabeçalhos duplicados)."""

    indices: dict[str, list[int]] = {}
    for posicao, nome in enumerate(cabecalho_normalizado):
        if nome:
            indices.setdefault(nome, []).append(posicao)
    return indices


def localizar(indices: dict[str, list[int]], *aliases: str, ocorrencia: int = 0) -> int | None:
    """Posição da coluna pelo primeiro alias presente (leitura adaptativa)."""

    for alias in aliases:
        posicoes = indices.get(alias)
        if posicoes and len(posicoes) > ocorrencia:
            return posicoes[ocorrencia]
    return None


def celula(linha: tuple[Any, ...], posicao: int | None) -> Any:
    if posicao is None or posicao >= len(linha):
        return None
    return linha[posicao]
